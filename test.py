# pyqt_main.py

import sys
import os
import json
import threading
import logging
import pandas as pd
import subprocess
import webbrowser

from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QLabel, QPushButton, QTabWidget, QLineEdit, QTextEdit, QCheckBox,
    QProgressBar, QListWidget, QListWidgetItem, QStyle, QInputDialog,
    QMessageBox
)
from PyQt5.QtCore import Qt, QThread, QSize
from PyQt5.QtGui import QIcon, QCursor

import pyperclip
import openpyxl

# Your existing modules
from scraper import run_scraper
from uploader import run_uploader

# Configure logging
logging.basicConfig(
    filename=os.path.join(os.getcwd(), 'app.log'),
    level=logging.INFO,
    format='%(asctime)s:%(levelname)s:%(message)s'
)

CONFIG_FILE = 'config.json'
EXCEL_FILE = 'scraped_data.xlsx'


def resource_path(relative_path):
    """
    Retrieves absolute path for resources, suitable for dev and PyInstaller.
    """
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)


def get_user_data_dir():
    """Determine directory for user data/config."""
    if getattr(sys, 'frozen', False):
        # If running as a PyInstaller bundle
        return os.path.dirname(sys.executable)
    else:
        return os.path.abspath(".")


def flatten_json(y):
    out = {}

    def flatten(x, name=''):
        if isinstance(x, dict):
            for a in x:
                if a == 'images':  # Skip images
                    continue
                flatten(x[a], f'{a}_' if name == '' else f'{name}{a}_')
        else:
            out[name[:-1]] = x

    flatten(y)
    return out


class RealEstateApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("ESTAGE")
        self.setGeometry(100, 100, 800, 700)

        try:
            self.setWindowIcon(QIcon(resource_path('logo.ico')))
        except Exception as e:
            logging.error(f"Failed to set icon: {e}")

        # State
        self.user_data_dir = get_user_data_dir()
        self.thread_stop_event = threading.Event()
        self.all_ad_ids = []

        # PyQt widgets for "Scrape & Upload" tab
        self.url_input = None
        self.agency_price_input = None
        self.comment_input = None
        self.headless_checkbox_scrape = None
        self.upload_desc_checkbox_scrape = None
        self.progress_scrape_upload = None
        self.run_button_scrape_upload = None
        self.stop_button_scrape_upload = None

        # PyQt widgets for "Scrape Only" tab
        self.url_input_scrape_only = None
        self.agency_price_input_scrape_only = None
        self.comment_input_scrape_only = None
        self.headless_checkbox_scrape_only = None
        self.progress_scrape_only = None
        self.run_button_scrape_only = None
        self.stop_button_scrape_only = None

        # PyQt widgets for "Upload Existing" tab
        self.ad_id_input = None
        self.ad_id_listwidget = None
        self.upload_desc_checkbox_upload = None
        self.progress_upload_existing = None
        self.run_button_upload_existing = None
        self.stop_button_upload_existing = None
        self.upload_link_label = None
        self.copy_url_button = None

        # Config and UI Setup
        self.ensure_user_data_dir_exists()
        self.user_config = self.load_or_create_config()
        self.init_ui()

    # ---------------------------
    #  Directory / Config Methods
    # ---------------------------
    def ensure_user_data_dir_exists(self):
        if not os.path.exists(self.user_data_dir):
            try:
                os.makedirs(self.user_data_dir)
                logging.info("User data directory created.")
            except Exception as e:
                logging.error(f"Failed to create user data directory: {e}")
                QMessageBox.critical(
                    self, "Folder Error",
                    f"Failed to create user data directory: {e}"
                )

    def load_or_create_config(self):
        config_path = os.path.join(self.user_data_dir, CONFIG_FILE)
        if os.path.exists(config_path):
            try:
                with open(config_path, 'r', encoding='utf-8') as f:
                    cfg = json.load(f)
                if 'email' in cfg and 'password' in cfg and 'name' in cfg:
                    logging.info("Configuration loaded successfully.")
                    return cfg
                else:
                    raise json.decoder.JSONDecodeError("Missing fields", "", 0)
            except json.decoder.JSONDecodeError:
                QMessageBox.critical(
                    self,
                    "Configuration Error",
                    "config.json is corrupted or missing required fields. Please re-enter your credentials."
                )
                os.remove(config_path)
                return self.show_login_dialog()
            except Exception as e:
                logging.error(f"Error loading config: {e}")
                QMessageBox.critical(
                    self,
                    "Configuration Error",
                    f"An error occurred while loading config.json: {e}"
                )
                return self.show_login_dialog()
        else:
            return self.show_login_dialog()

    def show_login_dialog(self):
        # Simple login dialog to capture user credentials
        email, ok1 = QInputDialog.getText(
            self, "User Login", "Enter email:"
        )
        if not ok1 or not email.strip():
            return None
        password, ok2 = QInputDialog.getText(
            self, "User Login", "Enter password:", echo=QLineEdit.Password
        )
        if not ok2 or not password.strip():
            return None
        name, ok3 = QInputDialog.getText(
            self, "User Login", "Enter name:"
        )
        if not ok3 or not name.strip():
            return None

        cfg = {
            'email': email.strip(),
            'password': password.strip(),
            'name': name.strip()
        }
        try:
            with open(os.path.join(self.user_data_dir, CONFIG_FILE), 'w', encoding='utf-8') as f:
                json.dump(cfg, f, indent=4, ensure_ascii=False)
            logging.info("Configuration saved successfully.")
        except Exception as e:
            logging.error(f"Failed to save configuration: {e}")
            QMessageBox.critical(
                self,
                "File Error",
                f"Failed to save configuration: {e}"
            )
            return None
        return cfg

    # -------------
    #  UI Building
    # -------------
    def init_ui(self):
        """Build the main UI depending on whether user_config is available."""
        if not self.user_config:
            # If config is missing or the user canceled the login, close the app
            self.close()
            return

        self.build_main_window()
        self.apply_global_styles()

    def build_main_window(self):
        # Create a central widget to hold everything
        central_widget = QWidget()
        main_layout = QVBoxLayout()
        central_widget.setLayout(main_layout)
        self.setCentralWidget(central_widget)

        # Header
        header_layout = QHBoxLayout()
        label_estage = QLabel("ESTAGE")
        label_estage.setStyleSheet("font-size: 24px; font-weight: bold;")
        header_layout.addWidget(label_estage)
        header_layout.addStretch()

        user_name = self.user_config.get('name', 'User')
        label_user = QLabel(f"Logged in as: {user_name}")
        label_user.setStyleSheet("font-size: 14px;")
        header_layout.addWidget(label_user)
        main_layout.addLayout(header_layout)

        # Notebook (Tabs)
        tabs = QTabWidget()
        main_layout.addWidget(tabs)

        # Initialize tab widgets
        scrape_upload_tab = QWidget()
        scrape_only_tab = QWidget()
        upload_existing_tab = QWidget()

        tabs.addTab(scrape_upload_tab, "Scrape & Upload")
        tabs.addTab(scrape_only_tab, "Scrape")
        tabs.addTab(upload_existing_tab, "Upload Existing")

        # Build "Scrape & Upload" Tab
        self.build_scrape_upload_tab(scrape_upload_tab)

        # Build "Scrape Only" Tab
        self.build_scrape_only_tab(scrape_only_tab)

        # Build "Upload Existing" Tab
        self.build_upload_existing_tab(upload_existing_tab)

        # Bottom buttons
        button_layout = QHBoxLayout()
        main_layout.addLayout(button_layout)

        btn_change_user = QPushButton("Change User")
        btn_change_user.setObjectName("btnDanger")
        btn_change_user.clicked.connect(self.change_user)
        button_layout.addWidget(btn_change_user)

        btn_open_excel = QPushButton("Open Excel")
        btn_open_excel.setObjectName("btnInfo")
        btn_open_excel.clicked.connect(self.open_excel_file)
        button_layout.addWidget(btn_open_excel)

        button_layout.addStretch()

        # Ensure data folder / excel file
        self.ensure_data_folder_exists()
        self.ensure_excel_file_exists()

        # Load Ad IDs
        self.all_ad_ids = self.get_all_ad_ids()

    def build_scrape_upload_tab(self, parent_widget):
        layout = QVBoxLayout()
        parent_widget.setLayout(layout)

        # URL Input
        label_url = QLabel("Enter URL:")
        layout.addWidget(label_url)
        self.url_input = QLineEdit()
        layout.addWidget(self.url_input)

        # Agency Price
        label_price = QLabel("Agency Price:")
        layout.addWidget(label_price)
        self.agency_price_input = QLineEdit()
        layout.addWidget(self.agency_price_input)

        # Comment
        label_comment = QLabel("Comment:")
        layout.addWidget(label_comment)
        self.comment_input = QLineEdit()
        layout.addWidget(self.comment_input)

        # Upload Description Checkbox
        self.upload_desc_checkbox_scrape = QCheckBox("Upload Description")
        self.upload_desc_checkbox_scrape.setChecked(True)
        layout.addWidget(self.upload_desc_checkbox_scrape)

        # Headless
        self.headless_checkbox_scrape = QCheckBox("Run scraping in headless mode")
        self.headless_checkbox_scrape.setChecked(True)
        layout.addWidget(self.headless_checkbox_scrape)

        # Progress Bar
        self.progress_scrape_upload = QProgressBar()
        self.progress_scrape_upload.setVisible(False)
        layout.addWidget(self.progress_scrape_upload)

        # Buttons
        btn_run = QPushButton("Run")
        btn_run.setObjectName("btnSuccess")
        btn_run.clicked.connect(self.start_scrape_upload)
        layout.addWidget(btn_run)

        btn_stop = QPushButton("Stop")
        btn_stop.setObjectName("btnDanger")
        btn_stop.clicked.connect(self.stop_running_process)
        btn_stop.setVisible(False)  # We can show/hide on run
        layout.addWidget(btn_stop)

        self.run_button_scrape_upload = btn_run
        self.stop_button_scrape_upload = btn_stop

    def build_scrape_only_tab(self, parent_widget):
        layout = QVBoxLayout()
        parent_widget.setLayout(layout)

        label_url = QLabel("Enter URL:")
        layout.addWidget(label_url)
        self.url_input_scrape_only = QLineEdit()
        layout.addWidget(self.url_input_scrape_only)

        label_price = QLabel("Agency Price:")
        layout.addWidget(label_price)
        self.agency_price_input_scrape_only = QLineEdit()
        layout.addWidget(self.agency_price_input_scrape_only)

        label_comment = QLabel("Comment:")
        layout.addWidget(label_comment)
        self.comment_input_scrape_only = QLineEdit()
        layout.addWidget(self.comment_input_scrape_only)

        self.headless_checkbox_scrape_only = QCheckBox("Run in headless mode")
        self.headless_checkbox_scrape_only.setChecked(True)
        layout.addWidget(self.headless_checkbox_scrape_only)

        self.progress_scrape_only = QProgressBar()
        self.progress_scrape_only.setVisible(False)
        layout.addWidget(self.progress_scrape_only)

        btn_scrape = QPushButton("Scrape")
        btn_scrape.setObjectName("btnSuccess")
        btn_scrape.clicked.connect(self.start_scrape_only)
        layout.addWidget(btn_scrape)

        btn_stop = QPushButton("Stop")
        btn_stop.setObjectName("btnDanger")
        btn_stop.clicked.connect(self.stop_running_process)
        btn_stop.setVisible(False)
        layout.addWidget(btn_stop)

        self.run_button_scrape_only = btn_scrape
        self.stop_button_scrape_only = btn_stop

    def build_upload_existing_tab(self, parent_widget):
        layout = QVBoxLayout()
        parent_widget.setLayout(layout)

        label_ad_id = QLabel("Enter Ad ID:")
        layout.addWidget(label_ad_id)
        self.ad_id_input = QLineEdit()
        self.ad_id_input.textChanged.connect(self.update_ad_id_autocomplete)
        layout.addWidget(self.ad_id_input)

        self.ad_id_listwidget = QListWidget()
        self.ad_id_listwidget.itemClicked.connect(self.on_ad_id_select)
        self.ad_id_listwidget.setVisible(False)
        layout.addWidget(self.ad_id_listwidget)

        self.upload_desc_checkbox_upload = QCheckBox("Upload Description")
        self.upload_desc_checkbox_upload.setChecked(True)
        layout.addWidget(self.upload_desc_checkbox_upload)

        self.progress_upload_existing = QProgressBar()
        self.progress_upload_existing.setVisible(False)
        layout.addWidget(self.progress_upload_existing)

        btn_upload = QPushButton("Upload")
        btn_upload.setObjectName("btnSuccess")
        btn_upload.clicked.connect(self.start_upload_existing)
        layout.addWidget(btn_upload)

        btn_stop = QPushButton("Stop")
        btn_stop.setObjectName("btnDanger")
        btn_stop.clicked.connect(self.stop_running_process)
        btn_stop.setVisible(False)
        layout.addWidget(btn_stop)

        self.run_button_upload_existing = btn_upload
        self.stop_button_upload_existing = btn_stop

        self.upload_link_label = QLabel()
        self.upload_link_label.setStyleSheet("color: #007bff; text-decoration: underline;")
        self.upload_link_label.setCursor(QCursor(Qt.PointingHandCursor))
        layout.addWidget(self.upload_link_label)

        self.copy_url_button = QPushButton("Copy URL")
        self.copy_url_button.setObjectName("btnInfo")
        self.copy_url_button.setEnabled(False)
        self.copy_url_button.clicked.connect(self.copy_url)
        layout.addWidget(self.copy_url_button)

    # -----------------
    #  Data Prep/Checks
    # -----------------
    def ensure_data_folder_exists(self):
        data_folder = os.path.join(self.user_data_dir, 'data')
        if not os.path.exists(data_folder):
            try:
                os.makedirs(data_folder)
                logging.info("Data folder created.")
            except Exception as e:
                logging.error(f"Failed to create data folder: {e}")
                QMessageBox.critical(
                    self, "Folder Error",
                    f"Failed to create data folder: {e}"
                )

    def ensure_excel_file_exists(self):
        excel_path = os.path.join(self.user_data_dir, EXCEL_FILE)
        if not os.path.exists(excel_path):
            columns = [
                "Ad ID", "Ad Title", "Location", "Number",
                "Owner Price", "Agency Price", "Phone Number", "Name",
                "Description", "Comment", "საერთო ფართი", "ოთახი",
                "საძინებელი", "სართული", "სართულიანობა", "სველი წერტილი",
                "მდგომარეობა", "სტატუსი", "category", "property_type",
                "transaction_type", "კონდიციონერი", "აივანი", "სარდაფი",
                "საკაბელო ტელევიზია", "ლიფტი", "მაცივარი", "ავეჯი",
                "გარაჟი", "მინა-პაკეტი", "ცენტ. გათბობა", "ცხელი წყალი",
                "ინტერნეტი", "რკინის კარი", "ბუნებრივი აირი", "სიგნალიზაცია",
                "სათავსო", "ტელეფონი", "ტელევიზორი", "სარეცხი მანქანა",
                "Final URL"
            ]
            try:
                df = pd.DataFrame(columns=columns)
                df.to_excel(excel_path, index=False)
                logging.info("Excel file created with proper columns.")
            except Exception as e:
                logging.error(f"Failed to create Excel file: {e}")
                QMessageBox.critical(
                    self, "Excel Error",
                    f"Failed to create Excel file: {e}"
                )

    def get_all_ad_ids(self):
        data_folder = os.path.join(self.user_data_dir, 'data')
        if not os.path.exists(data_folder):
            return []
        return [
            name for name in os.listdir(data_folder)
            if os.path.isdir(os.path.join(data_folder, name))
        ]

    # ---------------------------
    #  Scrape & Upload Tab Actions
    # ---------------------------
    def start_scrape_upload(self):
        url = self.url_input.text().strip()
        price = self.agency_price_input.text().strip()
        if not url or not price:
            QMessageBox.warning(self, "Input Error", "Please enter both URL and Agency Price.")
            return

        self.run_button_scrape_upload.setEnabled(False)
        self.stop_button_scrape_upload.setVisible(True)

        t = threading.Thread(target=self.run_scrape_upload, daemon=True)
        t.start()

    def run_scrape_upload(self):
        self.show_progress(self.progress_scrape_upload, True)
        try:
            headless = self.headless_checkbox_scrape.isChecked()
            upload_description = self.upload_desc_checkbox_scrape.isChecked()
            data_dir = os.path.join(self.user_data_dir, 'data')
            logging.info(f"Running scraper with data directory: {data_dir}")

            if self.thread_stop_event.is_set():
                self.show_info("Process was stopped.")
                return

            # Scrape
            ad_id = run_scraper(
                self.url_input.text().strip(),
                self.agency_price_input.text().strip(),
                comment=self.comment_input.text().strip(),
                headless=headless,
                stop_event=self.thread_stop_event,
                output_dir=data_dir
            )
            logging.info(f"Scraper returned Ad ID: {ad_id}")
            if self.thread_stop_event.is_set():
                self.show_info("Process was stopped.")
                return

            if not ad_id:
                self.show_error("Scraping failed. Check the URL and try again.")
                return

            # Check JSON existence
            json_file_path = os.path.join(data_dir, ad_id, f"{ad_id}.json")
            if not os.path.exists(json_file_path):
                logging.error("Scraped data file not found.")
                self.show_error("Scraped data file not found.")
                return

            # Load JSON
            with open(json_file_path, "r", encoding='utf-8') as jf:
                try:
                    scraped_data = json.load(jf)
                except json.JSONDecodeError as e:
                    logging.error(f"JSON decode error: {e}")
                    self.show_error(f"Failed to decode JSON file: {e}")
                    return

            flattened_data = flatten_json(scraped_data)
            # minimal example of excel_data
            excel_data = {
                "Ad ID": flattened_data.get("ad_id", ""),
                "Ad Title": flattened_data.get("ad_title", ""),
                "Owner Price": flattened_data.get("owner_price", ""),
                "Agency Price": flattened_data.get("agency_price", ""),
                "Description": flattened_data.get("description", ""),
                "Comment": flattened_data.get("comment", ""),
                "Final URL": ""
            }

            # Write to Excel
            excel_path = os.path.join(self.user_data_dir, EXCEL_FILE)
            try:
                if not os.path.exists(excel_path):
                    df = pd.DataFrame([excel_data])
                    df.to_excel(excel_path, index=False)
                    logging.info("Excel file created and data written successfully.")
                else:
                    df = pd.read_excel(excel_path)
                    df = pd.concat([df, pd.DataFrame([excel_data])], ignore_index=True)
                    df.to_excel(excel_path, index=False)
                    logging.info(f"Data appended to Excel for Ad ID: {ad_id}")
            except Exception as e:
                logging.error(f"Failed to write to Excel: {e}")
                self.show_error(f"Failed to write to Excel: {e}")
                return

            if self.thread_stop_event.is_set():
                self.show_info("Process was stopped.")
                return

            # Uploader
            final_url = run_uploader(
                username=self.user_config['email'],
                password=self.user_config['password'],
                phone_number=flattened_data.get("phone_number", ""),
                ad_id=ad_id,
                enter_description=upload_description,
                headless=False,
                stop_event=self.thread_stop_event,
                output_dir=data_dir
            )
            logging.info(f"Uploader returned Final URL: {final_url}")
            if self.thread_stop_event.is_set():
                self.show_info("Process was stopped.")
                return

            if final_url:
                # Update Final URL in Excel
                wb = openpyxl.load_workbook(excel_path)
                ws = wb.active
                header_map = {cell.value: idx for idx, cell in enumerate(ws[1], 1)}
                ad_id_col = header_map.get('Ad ID')
                final_url_col = header_map.get('Final URL')

                if not ad_id_col:
                    raise ValueError("'Ad ID' column not found in Excel.")
                if not final_url_col:
                    final_url_col = ws.max_column + 1
                    ws.cell(row=1, column=final_url_col, value='Final URL')

                updated = False
                for row in ws.iter_rows(min_row=2, values_only=False):
                    cell = row[ad_id_col - 1]
                    if cell.value == ad_id:
                        row[final_url_col - 1].value = final_url
                        updated = True
                        break

                if not updated:
                    raise ValueError(f"Ad ID {ad_id} not found in Excel.")

                wb.save(excel_path)
                logging.info(f"Final URL updated in Excel for Ad ID: {ad_id}")

                self.show_info("Scraping completed successfully and data saved to Excel.")
                self.upload_link_label.setText(f"Upload Successful!\nFinal URL: {final_url}")
                self.copy_url_button.setEnabled(True)
            else:
                self.show_error("Upload failed. Please check the logs for more details.")

            self.all_ad_ids = self.get_all_ad_ids()
        except Exception as e:
            logging.error(f"An error occurred in run_scrape_upload: {e}", exc_info=True)
            self.show_error(f"An error occurred: {e}")
        finally:
            self.show_progress(self.progress_scrape_upload, False)
            self.run_button_scrape_upload.setEnabled(True)
            self.stop_button_scrape_upload.setVisible(False)
            self.thread_stop_event.clear()

    # -----------------
    #  Scrape Only Tab
    # -----------------
    def start_scrape_only(self):
        if not self.url_input_scrape_only.text().strip() or not self.agency_price_input_scrape_only.text().strip():
            QMessageBox.warning(self, "Input Error", "Please enter both URL and Agency Price.")
            return

        self.run_button_scrape_only.setEnabled(False)
        self.stop_button_scrape_only.setVisible(True)
        t = threading.Thread(target=self.run_scrape_only_thread, daemon=True)
        t.start()

    def run_scrape_only_thread(self):
        self.show_progress(self.progress_scrape_only, True)
        try:
            if self.thread_stop_event.is_set():
                self.show_info("Process was stopped.")
                return
            data_dir = os.path.join(self.user_data_dir, 'data')
            logging.info(f"Running scraper with data directory: {data_dir}")

            headless = self.headless_checkbox_scrape_only.isChecked()
            ad_id = run_scraper(
                self.url_input_scrape_only.text().strip(),
                self.agency_price_input_scrape_only.text().strip(),
                comment=self.comment_input_scrape_only.text().strip(),
                headless=headless,
                stop_event=self.thread_stop_event,
                output_dir=data_dir
            )
            logging.info(f"Scraper returned Ad ID: {ad_id}")
            if self.thread_stop_event.is_set():
                self.show_info("Process was stopped.")
                return

            if not ad_id:
                self.show_error("Scraping failed. Check the URL and try again.")
                return

            json_file_path = os.path.join(data_dir, ad_id, f"{ad_id}.json")
            if not os.path.exists(json_file_path):
                logging.error("Scraped data file not found.")
                self.show_error("Scraped data file not found.")
                return

            with open(json_file_path, "r", encoding='utf-8') as jf:
                try:
                    scraped_data = json.load(jf)
                    logging.info("Scraped data loaded successfully.")
                except json.JSONDecodeError as e:
                    logging.error(f"JSON decode error: {e}")
                    self.show_error(f"Failed to decode JSON file: {e}")
                    return

            flattened_data = flatten_json(scraped_data)
            excel_data = {
                "Ad ID": flattened_data.get("ad_id", ""),
                "Ad Title": flattened_data.get("ad_title", ""),
                "Final URL": ""
            }

            excel_path = os.path.join(self.user_data_dir, EXCEL_FILE)
            try:
                if not os.path.exists(excel_path):
                    df = pd.DataFrame([excel_data])
                    df.to_excel(excel_path, index=False)
                    logging.info("Excel file created and data written successfully.")
                else:
                    df = pd.read_excel(excel_path)
                    df = pd.concat([df, pd.DataFrame([excel_data])], ignore_index=True)
                    df.to_excel(excel_path, index=False)
                    logging.info(f"Data appended to Excel for Ad ID: {ad_id}")
            except Exception as e:
                logging.error(f"Failed to write to Excel: {e}")
                self.show_error(f"Failed to write to Excel: {e}")
                return

            self.show_info("Scraping completed successfully and data saved to Excel.")
        except Exception as e:
            logging.error(f"An error occurred in run_scrape_only: {e}", exc_info=True)
            self.show_error(f"An error occurred: {e}")
        finally:
            self.show_progress(self.progress_scrape_only, False)
            self.run_button_scrape_only.setEnabled(True)
            self.stop_button_scrape_only.setVisible(False)
            self.thread_stop_event.clear()

    # -------------------
    #  Upload Existing Tab
    # -------------------
    def start_upload_existing(self):
        ad_id_val = self.ad_id_input.text().strip()
        if not ad_id_val:
            QMessageBox.warning(self, "Input Error", "Please enter an Ad ID.")
            return

        self.run_button_upload_existing.setEnabled(False)
        self.stop_button_upload_existing.setVisible(True)
        t = threading.Thread(target=self.run_upload_existing_thread, args=(ad_id_val,), daemon=True)
        t.start()

    def run_upload_existing_thread(self, ad_id_val):
        self.show_progress(self.progress_upload_existing, True)
        try:
            excel_path = os.path.join(self.user_data_dir, EXCEL_FILE)
            upload_description = self.upload_desc_checkbox_upload.isChecked()

            if self.thread_stop_event.is_set():
                self.show_info("Process was stopped.")
                return

            json_file_path = os.path.join(self.user_data_dir, 'data', ad_id_val, f"{ad_id_val}.json")
            if not os.path.exists(json_file_path):
                logging.error("Scraped data file not found.")
                self.show_error("Scraped data file not found.")
                return

            with open(json_file_path, "r", encoding='utf-8') as jf:
                try:
                    scraped_data = json.load(jf)
                except json.JSONDecodeError as e:
                    logging.error(f"JSON decode error: {e}")
                    self.show_error(f"Failed to decode JSON file: {e}")
                    return

            final_url = run_uploader(
                username=self.user_config['email'],
                password=self.user_config['password'],
                phone_number=scraped_data.get("phone_number", ""),
                ad_id=ad_id_val,
                enter_description=upload_description,
                headless=False,
                stop_event=self.thread_stop_event,
                output_dir=os.path.join(self.user_data_dir, 'data')
            )
            logging.info(f"Uploader returned Final URL: {final_url}")

            if self.thread_stop_event.is_set():
                self.show_info("Process was stopped.")
                return

            if final_url:
                wb = openpyxl.load_workbook(excel_path)
                ws = wb.active
                header_map = {cell.value: idx for idx, cell in enumerate(ws[1], 1)}
                ad_id_col = header_map.get('Ad ID')
                final_url_col = header_map.get('Final URL')

                if not ad_id_col:
                    raise ValueError("'Ad ID' column not found in Excel.")
                if not final_url_col:
                    final_url_col = ws.max_column + 1
                    ws.cell(row=1, column=final_url_col, value='Final URL')

                updated = False
                for row in ws.iter_rows(min_row=2, values_only=False):
                    if row[ad_id_col - 1].value == ad_id_val:
                        row[final_url_col - 1].value = final_url
                        updated = True
                        break

                if not updated:
                    raise ValueError(f"Ad ID {ad_id_val} not found in Excel.")

                wb.save(excel_path)
                logging.info(f"Final URL updated in Excel for Ad ID: {ad_id_val}")

                self.show_info("Scraping completed successfully and data saved to Excel.")
                self.upload_link_label.setText(f"Upload Successful!\nFinal URL: {final_url}")
                self.copy_url_button.setEnabled(True)
            else:
                self.show_error("Upload failed. Please check the logs for more details.")
        except Exception as e:
            logging.error(f"An error occurred in run_upload_existing: {e}", exc_info=True)
            self.show_error(f"An error occurred: {e}")
        finally:
            self.show_progress(self.progress_upload_existing, False)
            self.run_button_upload_existing.setEnabled(True)
            self.stop_button_upload_existing.setVisible(False)
            self.thread_stop_event.clear()

    # -------------------
    #  Shared Functions
    # -------------------
    def apply_global_styles(self):
        """
        Apply a custom QSS style that mimics a flat Bootstrap-inspired design.
        Adjust colors, border radii, paddings, etc. to suit your preference.
        """
        qss = """
        /* Main Window */
        QMainWindow {
            background-color: #f8f9fa;
        }

        /* Tabs */
        QTabWidget::pane {
            border: 1px solid #dee2e6;
            background-color: #ffffff;
        }
        QTabBar::tab {
            background-color: #e9ecef;
            border: 1px solid #dee2e6;
            padding: 6px 12px;
            margin: 0;
        }
        QTabBar::tab:selected {
            background-color: #ffffff;
            border-bottom: none;
        }

        /* Buttons */
        QPushButton {
            border: none;
            border-radius: 4px;
            padding: 6px 12px;
            font-size: 14px;
            color: #ffffff;
        }
        QPushButton#btnSuccess {
            background-color: #28a745;
        }
        QPushButton#btnSuccess:hover {
            background-color: #218838;
        }
        QPushButton#btnDanger {
            background-color: #dc3545;
        }
        QPushButton#btnDanger:hover {
            background-color: #c82333;
        }
        QPushButton#btnInfo {
            background-color: #17a2b8;
        }
        QPushButton#btnInfo:hover {
            background-color: #117a8b;
        }
        QPushButton:disabled {
            background-color: #6c757d;
            color: #ffffff;
        }

        /* Labels */
        QLabel {
            font-size: 14px;
            color: #212529;
        }

        /* QLineEdit, QCheckBox, etc. */
        QLineEdit, QTextEdit {
            background-color: #ffffff;
            border: 1px solid #ced4da;
            border-radius: 4px;
            padding: 4px;
            font-size: 14px;
        }
        QCheckBox {
            font-size: 14px;
            spacing: 6px;
        }
        QCheckBox::indicator {
            width: 18px;
            height: 18px;
        }

        /* Progress Bars */
        QProgressBar {
            border: 1px solid #ced4da;
            border-radius: 4px;
            background-color: #e9ecef;
            text-align: center;
        }
        QProgressBar::chunk {
            background-color: #007bff;
        }

        /* QListWidget */
        QListWidget {
            border: 1px solid #ced4da;
            border-radius: 4px;
            font-size: 14px;
            background-color: #ffffff;
        }

        /* Message Boxes, etc. inherit default fonts/colors */
        """
        self.setStyleSheet(qss)

    def show_progress(self, progress_bar, visible):
        """Helper to show/hide a QProgressBar with indefinite mode."""
        progress_bar.setVisible(visible)
        if visible:
            progress_bar.setRange(0, 0)  # indefinite
        else:
            progress_bar.setRange(0, 1)  # reset

    def stop_running_process(self):
        self.thread_stop_event.set()
        logging.info("Stop event triggered.")

    def show_error(self, message):
        QMessageBox.critical(self, "Error", message)

    def show_info(self, message):
        QMessageBox.information(self, "Success", message)

    def open_excel_file(self):
        excel_path = os.path.join(self.user_data_dir, EXCEL_FILE)
        if not os.path.exists(excel_path):
            self.show_error("Excel file does not exist.")
            return
        try:
            if sys.platform.startswith('darwin'):
                subprocess.call(('open', excel_path))
            elif os.name == 'nt':
                os.startfile(excel_path)
            elif os.name == 'posix':
                subprocess.call(('xdg-open', excel_path))
        except Exception as e:
            logging.error(f"Failed to open Excel file: {e}")
            self.show_error(f"Failed to open Excel file: {e}")

    # Our label is clickable (mousePressEvent)
    def mousePressEvent(self, event):
        # We override QMainWindow's mousePressEvent, but it's optional.
        pass

    def on_ad_id_select(self, item):
        if item:
            self.ad_id_input.setText(item.text())
            self.ad_id_listwidget.setVisible(False)

    def update_ad_id_autocomplete(self, text):
        typed = text.strip().lower()
        if not typed:
            data = self.all_ad_ids
        else:
            data = [ad_id for ad_id in self.all_ad_ids if typed in ad_id.lower()]

        self.ad_id_listwidget.clear()
        if data:
            for item in data:
                self.ad_id_listwidget.addItem(item)
            self.ad_id_listwidget.setVisible(True)
        else:
            self.ad_id_listwidget.setVisible(False)

    def copy_url(self):
        text = self.upload_link_label.text()
        if "Final URL:" in text:
            final_url = text.split("Final URL: ")[-1].strip()
            if final_url:
                pyperclip.copy(final_url)
                QMessageBox.information(self, "Copied", "URL has been copied to clipboard.")

    def change_user(self):
        confirm = QMessageBox.question(
            self, "Change User", "Are you sure you want to change the user?",
            QMessageBox.Yes | QMessageBox.No
        )
        if confirm == QMessageBox.Yes:
            config_path = os.path.join(self.user_data_dir, CONFIG_FILE)
            if os.path.exists(config_path):
                try:
                    os.remove(config_path)
                    logging.info("Configuration file deleted.")
                except Exception as e:
                    logging.error(f"Failed to delete config.json: {e}")
                    QMessageBox.critical(self, "File Error", f"Failed to delete config.json: {e}")
                    return
            self.close() 
    # -------------
    #  Main Entry
    # -------------
def main():
    app = QApplication(sys.argv)
    window = RealEstateApp()
    window.show()
    sys.exit(app.exec_())


if __name__ == "__main__":
    main()
