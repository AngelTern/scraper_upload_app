# main.py

import datetime
import openpyxl
import pyperclip
import ttkbootstrap as ttk
from ttkbootstrap.constants import *
from scraper import run_scraper
from uploader import run_uploader
from threading import Thread, Event
import os
import json
import tkinter as tk
from tkinter import messagebox
import sys
import pandas as pd
import subprocess
import webbrowser
import logging

# Configure logging
logging.basicConfig(
    filename=os.path.join(os.getcwd(), 'app.log'),
    level=logging.INFO,
    format='%(asctime)s:%(levelname)s:%(message)s'
)

CONFIG_FILE = 'config.json'
EXCEL_FILE = 'scraped_data.xlsx'

def get_user_data_dir():
    """
    Determines the directory in which to place user data (config, Excel, etc.).
    """
    if getattr(sys, 'frozen', False):
        return os.path.dirname(sys.executable)
    else:
        return os.path.abspath(".")

def resource_path(relative_path):
    """
    Safely get absolute path to resources, supporting PyInstaller's _MEIPASS.
    """
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)

def flatten_json(y):
    """
    Flatten a nested JSON/dict structure into a single dict with 'dot-like' keys.
    """
    out = {}
    def flatten(x, name=''):
        if isinstance(x, dict):
            for a in x:
                if a == 'images':  # Skip images
                    continue
                flatten(x[a], f'{a}_' if name == '' else f'{name}{a}_')
        else:
            out[name[:-1]] = x
    flatten(y)
    return out

class RealEstateApp:
    def __init__(self, root):
        self.root = root
        self.root.title("ESTAGE")
        self.root.geometry("800x700")

        try:
            self.root.iconbitmap(resource_path('logo.ico'))
        except Exception as e:
            logging.error(f"Failed to set icon: {e}")

        # Global state
        self.login_frame = None
        self.main_frame = None
        self.user_data_dir = get_user_data_dir()
        self.ensure_user_data_dir_exists()

        # Load or create config
        self.user_config = self.load_or_create_config()

        # Known Ad IDs in data folder
        self.all_ad_ids = self.get_all_ad_ids()

        # Thread stop event
        self.thread_stop_event = Event()

        # Build UI
        self.build_ui()

    def ensure_user_data_dir_exists(self):
        """
        Ensures the user data directory exists.
        """
        if not os.path.exists(self.user_data_dir):
            try:
                os.makedirs(self.user_data_dir)
                logging.info("User data directory created.")
            except Exception as e:
                logging.error(f"Failed to create user data directory: {e}")
                messagebox.showerror("Folder Error", f"Failed to create user data directory: {e}")

    def load_or_create_config(self):
        """
        Tries to load config.json. If missing or invalid, show login frame.
        """
        config_path = os.path.join(self.user_data_dir, CONFIG_FILE)
        if os.path.exists(config_path):
            try:
                with open(config_path, 'r', encoding='utf-8') as f:
                    config = json.load(f)
                if 'email' in config and 'password' in config and 'name' in config:
                    logging.info("Configuration loaded successfully.")
                    return config
                else:
                    raise json.decoder.JSONDecodeError("Missing fields", "", 0)
            except json.decoder.JSONDecodeError:
                messagebox.showerror(
                    "Configuration Error",
                    "config.json is corrupted or missing required fields. Please re-enter your credentials."
                )
                os.remove(config_path)
                return self.show_login_frame()
            except Exception as e:
                logging.error(f"Error loading config: {e}")
                messagebox.showerror(
                    "Configuration Error",
                    f"An error occurred while loading config.json: {e}"
                )
                return self.show_login_frame()
        else:
            return self.show_login_frame()

    def show_login_frame(self):
        """
        Show a login frame so the user can enter email/password/name
        and we create config.json from that.
        """
        if self.main_frame:
            self.main_frame.pack_forget()

        self.login_frame = ttk.Frame(self.root)
        self.login_frame.pack(expand=True, fill='both', pady=50)

        login_title = ttk.Label(
            self.login_frame,
            text="Please Log In",
            font=('Helvetica', 18, 'bold')
        )
        login_title.pack(pady=20)

        email_label = ttk.Label(self.login_frame, text="Email:")
        email_label.pack(pady=10, anchor='w', padx=20)
        self.email_entry = ttk.Entry(self.login_frame, width=40)
        self.email_entry.pack(pady=5)

        password_label = ttk.Label(self.login_frame, text="Password:")
        password_label.pack(pady=10, anchor='w', padx=20)
        self.password_entry = ttk.Entry(self.login_frame, width=40, show='*')
        self.password_entry.pack(pady=5)

        name_label = ttk.Label(self.login_frame, text="სახელი")
        name_label.pack(pady=10, anchor='w', padx=20)
        self.name_entry = ttk.Entry(self.login_frame, width=40)
        self.name_entry.pack(pady=5)

        login_button = ttk.Button(
            self.login_frame,
            text="Login",
            command=self.handle_login,
            style='primary.TButton'
        )
        login_button.pack(pady=20)
        return None

    def handle_login(self):
        """
        Called when user hits "Login" in the login frame. Creates config.json.
        """
        email = self.email_entry.get().strip()
        password = self.password_entry.get().strip()
        name = self.name_entry.get().strip()

        if not email or not password or not name:
            messagebox.showerror(
                "Input Error",
                "Please enter email, password, and name."
            )
            return

        config = {
            'email': email,
            'password': password,
            'name': name
        }
        try:
            with open(os.path.join(self.user_data_dir, CONFIG_FILE), 'w', encoding='utf-8') as f:
                json.dump(config, f, indent=4, ensure_ascii=False)
            logging.info("Configuration saved successfully.")
        except Exception as e:
            logging.error(f"Failed to save configuration: {e}")
            messagebox.showerror(
                "File Error",
                f"Failed to save configuration: {e}"
            )
            return

        self.user_config = config
        if self.login_frame:
            self.login_frame.pack_forget()
            self.login_frame.destroy()
            self.login_frame = None

        self.build_main_frame()

    def build_ui(self):
        """
        If config is valid, build the main frame; otherwise do nothing.
        """
        if self.user_config:
            self.build_main_frame()

    def build_main_frame(self):
        """
        Builds the main GUI with tabs: Scrape & Upload, Scrape, Upload Existing.
        Also ensures the Excel file and data folder exist.
        """
        self.ensure_data_folder_exists()
        self.ensure_excel_file_exists()

        self.main_frame = ttk.Frame(self.root)
        self.main_frame.pack(expand=True, fill='both', pady=10)

        header_frame = ttk.Frame(self.main_frame)
        header_frame.pack(fill='x')

        header_label = ttk.Label(
            header_frame,
            text="ESTAGE",
            style='primary.TLabel',
            font=('Helvetica', 24, 'bold')
        )
        header_label.pack(pady=20, side='left')

        user_name = self.user_config.get('name', 'User')
        user_label = ttk.Label(
            header_frame,
            text=f"Logged in as: {user_name}",
            style='secondary.TLabel',
            font=('Helvetica', 14)
        )
        user_label.pack(pady=20, side='right')

        # Notebook (tabs)
        self.notebook = ttk.Notebook(self.main_frame)
        self.notebook.pack(expand=True, fill='both', pady=10)

        self.scrape_upload_frame = ttk.Frame(self.notebook)
        self.scrape_only_frame = ttk.Frame(self.notebook)
        self.upload_existing_frame = ttk.Frame(self.notebook)

        self.notebook.add(self.scrape_upload_frame, text='Scrape & Upload')
        self.notebook.add(self.scrape_only_frame, text='Scrape')
        self.notebook.add(self.upload_existing_frame, text='Upload Existing')

        # Tkinter variables
        self.url = ttk.StringVar()
        self.agency_price = ttk.StringVar()
        self.comment = ttk.StringVar()  # user-typed comment
        self.headless_var_scrape = ttk.BooleanVar(value=True)
        self.upload_description_var_scrape = ttk.BooleanVar(value=True)
        self.existing_ad_id = ttk.StringVar()
        self.upload_description_var_upload = ttk.BooleanVar(value=True)
        self.upload_link = ttk.StringVar()

        self.build_scrape_upload_tab()
        self.build_scrape_only_tab()
        self.build_upload_existing_tab()

        change_user_button = ttk.Button(
            self.main_frame,
            text="Change User",
            command=self.change_user,
            style='danger.TButton'
        )
        change_user_button.pack(pady=10)

        open_excel_button = ttk.Button(
            header_frame,
            text="Open Excel",
            command=self.open_excel_file,
            style='info.TButton'
        )
        open_excel_button.pack(pady=20, side='right')

    def ensure_data_folder_exists(self):
        """
        Ensure 'data' folder exists in user_data_dir.
        """
        data_folder = os.path.join(self.user_data_dir, 'data')
        if not os.path.exists(data_folder):
            try:
                os.makedirs(data_folder)
                logging.info("Data folder created.")
            except Exception as e:
                logging.error(f"Failed to create data folder: {e}")
                messagebox.showerror(
                    "Folder Error",
                    f"Failed to create data folder: {e}"
                )

    def ensure_excel_file_exists(self):
        """
        Creates an Excel file with the columns:
        'Uploaded Timestamp', 'მესაკუთრის ID', 'ტელეფონის ნომერი',
        'ოთახი', 'სართული', 'მისამართი', 'სააგენტოს ფასი',
        'მესაკუთრის ფასი', 'Comment', 'ss.ge'
        if it doesn't already exist.
        """
        excel_path = os.path.join(self.user_data_dir, EXCEL_FILE)
        if not os.path.exists(excel_path):
            columns = [
                "Uploaded Timestamp",
                "მესაკუთრის ID",
                "ტელეფონის ნომერი",
                "ოთახი",
                "სართული",
                "მისამართი",
                "სააგენტოს ფასი",
                "მესაკუთრის ფასი",
                "Comment",
                "ss.ge"
            ]
            try:
                df = pd.DataFrame(columns=columns)
                df.to_excel(excel_path, index=False)
                logging.info("Excel file created with the necessary columns.")
            except Exception as e:
                logging.error(f"Failed to create Excel file: {e}")
                messagebox.showerror("Excel Error", f"Failed to create Excel file: {e}")

    def build_scrape_upload_tab(self):
        frame = self.scrape_upload_frame

        label_url = ttk.Label(frame, text="Enter URL:")
        label_url.pack(pady=5, anchor='w', padx=20)
        entry_url = ttk.Entry(frame, textvariable=self.url, width=60)
        entry_url.pack(pady=5, fill='x', padx=20)

        label_price = ttk.Label(frame, text="Agency Price:")
        label_price.pack(pady=5, anchor='w', padx=20)
        entry_price = ttk.Entry(frame, textvariable=self.agency_price, width=60)
        entry_price.pack(pady=5, fill='x', padx=20)

        label_comment = ttk.Label(frame, text="Comment:")
        label_comment.pack(pady=5, anchor='w', padx=20)
        entry_comment = ttk.Entry(frame, textvariable=self.comment, width=60)
        entry_comment.pack(pady=5, fill='x', padx=20)

        upload_desc_checkbox = ttk.Checkbutton(
            frame,
            text="Upload Description",
            variable=self.upload_description_var_scrape
        )
        upload_desc_checkbox.pack(pady=5, anchor='w', padx=20)

        checkbox_headless = ttk.Checkbutton(
            frame,
            text="Run scraping in headless mode",
            variable=self.headless_var_scrape
        )
        checkbox_headless.pack(pady=5, anchor='w', padx=20)

        self.progress_scrape_upload = ttk.Progressbar(frame, mode='indeterminate')
        self.progress_scrape_upload.pack(pady=5, fill='x', padx=20)
        self.progress_scrape_upload.pack_forget()

        self.run_button = ttk.Button(
            frame,
            text="Run",
            command=self.start_scrape_upload,
            style='success.TButton'
        )
        self.run_button.pack(pady=20)

        self.stop_button = ttk.Button(
            frame,
            text="Stop",
            command=self.stop_running_process,
            style='danger.TButton'
        )
        self.stop_button.pack(pady=5)
        self.stop_button.pack_forget()

    def build_scrape_only_tab(self):
        frame = self.scrape_only_frame

        label_url = ttk.Label(frame, text="Enter URL:")
        label_url.pack(pady=5, anchor='w', padx=20)
        entry_url = ttk.Entry(frame, textvariable=self.url, width=60)
        entry_url.pack(pady=5, fill='x', padx=20)

        label_price = ttk.Label(frame, text="Agency Price:")
        label_price.pack(pady=5, anchor='w', padx=20)
        entry_price = ttk.Entry(frame, textvariable=self.agency_price, width=60)
        entry_price.pack(pady=5, fill='x', padx=20)

        label_comment = ttk.Label(frame, text="Comment:")
        label_comment.pack(pady=5, anchor='w', padx=20)
        entry_comment = ttk.Entry(frame, textvariable=self.comment, width=60)
        entry_comment.pack(pady=5, fill='x', padx=20)

        checkbox_headless = ttk.Checkbutton(
            frame,
            text="Run in headless mode",
            variable=self.headless_var_scrape
        )
        checkbox_headless.pack(pady=5, anchor='w', padx=20)

        self.progress_scrape_only = ttk.Progressbar(frame, mode='indeterminate')
        self.progress_scrape_only.pack(pady=5, fill='x', padx=20)
        self.progress_scrape_only.pack_forget()

        self.run_button_scrape_only = ttk.Button(
            frame,
            text="Scrape",
            command=self.start_scrape_only,
            style='success.TButton'
        )
        self.run_button_scrape_only.pack(pady=20)

        self.stop_button_scrape_only = ttk.Button(
            frame,
            text="Stop",
            command=self.stop_running_process,
            style='danger.TButton'
        )
        self.stop_button_scrape_only.pack(pady=5)
        self.stop_button_scrape_only.pack_forget()

    def build_upload_existing_tab(self):
        frame = self.upload_existing_frame

        label_ad_id = ttk.Label(frame, text="Enter Ad ID:")
        label_ad_id.pack(pady=5, anchor='w', padx=20)
        self.ad_id_entry = ttk.Entry(frame, textvariable=self.existing_ad_id, width=60)
        self.ad_id_entry.pack(pady=5, fill='x', padx=20)
        self.ad_id_entry.bind('<KeyRelease>', self.update_ad_id_autocomplete)

        self.ad_id_listbox = tk.Listbox(frame, height=5)
        self.ad_id_listbox.pack(pady=5, fill='x', padx=20)
        self.ad_id_listbox.bind('<<ListboxSelect>>', self.on_ad_id_select)
        self.ad_id_listbox.pack_forget()

        upload_desc_checkbox = ttk.Checkbutton(
            frame,
            text="Upload Description",
            variable=self.upload_description_var_upload
        )
        upload_desc_checkbox.pack(pady=5, anchor='w', padx=20)

        self.progress_upload_existing = ttk.Progressbar(frame, mode='indeterminate')
        self.progress_upload_existing.pack(pady=5, fill='x', padx=20)
        self.progress_upload_existing.pack_forget()

        self.run_button_upload_existing = ttk.Button(
            frame,
            text="Upload",
            command=self.start_upload_existing,
            style='success.TButton'
        )
        self.run_button_upload_existing.pack(pady=20)

        self.stop_button_upload_existing = ttk.Button(
            frame,
            text="Stop",
            command=self.stop_running_process,
            style='danger.TButton'
        )
        self.stop_button_upload_existing.pack(pady=5)
        self.stop_button_upload_existing.pack_forget()

        self.final_url_label = ttk.Label(
            frame,
            textvariable=self.upload_link,
            foreground="blue",
            cursor="hand2",
            wraplength=700
        )
        self.final_url_label.pack(pady=10, padx=20)
        self.final_url_label.bind("<Button-1>", self.open_url)

        copy_button = ttk.Button(
            frame,
            text="Copy URL",
            command=self.copy_url,
            state='disabled',
            style='info.TButton'
        )
        copy_button.pack(pady=5)
        self.copy_button = copy_button

    def start_scrape_upload(self):
        """
        Validate inputs, disable run button, show stop button, run scraping in a thread.
        """
        if not self.validate_scrape_upload_inputs():
            return
        self.run_button.config(state='disabled')
        self.stop_button.pack(pady=5)
        Thread(target=self.run_scrape_upload, daemon=True).start()

    def run_scrape_upload(self):
        """
        1) Scrape data
        2) Write minimal fields to Excel, including user-typed comment
        3) Run uploader
        4) On success, set 'Uploaded Timestamp' and 'ss.ge' columns
        """
        try:
            self.progress_scrape_upload.pack(pady=5, fill='x', padx=20)
            self.progress_scrape_upload.start()

            headless = self.headless_var_scrape.get()
            upload_description = self.upload_description_var_scrape.get()

            if self.thread_stop_event.is_set():
                self.show_info("Process was stopped.")
                return

            data_dir = os.path.join(self.user_data_dir, 'data')
            logging.info(f"Running scraper with data directory: {data_dir}")

            ad_id = run_scraper(
                self.url.get(),
                self.agency_price.get(),
                comment=self.comment.get(),  # pass user comment to the scraper
                headless=headless,
                stop_event=self.thread_stop_event,
                output_dir=data_dir
            )
            logging.info(f"Scraper returned Ad ID: {ad_id}")

            if self.thread_stop_event.is_set():
                self.show_info("Process was stopped.")
                return

            if not ad_id:
                self.show_error("Scraping failed. Check the URL and try again.")
                return

            # Load the JSON data that scraper created
            json_file_path = os.path.join(data_dir, ad_id, f"{ad_id}.json")
            if not os.path.exists(json_file_path):
                logging.error("Scraped data file not found.")
                self.show_error("Scraped data file not found.")
                return

            with open(json_file_path, "r", encoding='utf-8') as jf:
                try:
                    scraped_data = json.load(jf)
                    logging.info("Scraped data loaded successfully.")
                except json.JSONDecodeError as e:
                    logging.error(f"JSON decode error: {e}")
                    self.show_error(f"Failed to decode JSON file: {e}")
                    return

            flattened_data = flatten_json(scraped_data)

            # Build Excel row (comment is from user, stored in self.comment)
            excel_data = {
                "Uploaded Timestamp": "",
                "მესაკუთრის ID": flattened_data.get("ad_id", ""),
                "ტელეფონის ნომერი": flattened_data.get("phone_number", ""),
                "ოთახი": flattened_data.get("property_details_ოთახი", ""),
                "სართული": flattened_data.get("property_details_სართული", ""),
                "მისამართი": f"{flattened_data.get('location', '')} {flattened_data.get('number', '')}".strip(),
                "სააგენტოს ფასი": flattened_data.get("agency_price", ""),
                "მესაკუთრის ფასი": flattened_data.get("owner_price", ""),
                "Comment": self.comment.get(),  # <-- user typed comment
                "ss.ge": ""
            }

            excel_path = os.path.join(self.user_data_dir, EXCEL_FILE)
            try:
                if not os.path.exists(excel_path):
                    self.ensure_excel_file_exists()

                df = pd.read_excel(excel_path)
                df = pd.concat([df, pd.DataFrame([excel_data])], ignore_index=True)
                df.to_excel(excel_path, index=False)
                logging.info(f"Data appended to Excel for Ad ID: {ad_id}")
            except Exception as e:
                logging.error(f"Failed to write to Excel: {e}")
                self.show_error(f"Failed to write to Excel: {e}")
                return

            if self.thread_stop_event.is_set():
                self.show_info("Process was stopped.")
                return

            # Uploader
            user_info = self.user_config
            final_url = run_uploader(
                username=user_info['email'],
                password=user_info['password'],
                phone_number=flattened_data.get("phone_number", ""),
                ad_id=ad_id,
                enter_description=upload_description,
                headless=False,  # forced false or set headless if you prefer
                stop_event=self.thread_stop_event,
                output_dir=data_dir
            )
            logging.info(f"Uploader returned Final URL: {final_url}")

            if self.thread_stop_event.is_set():
                self.show_info("Process was stopped.")
                return

            if final_url:
                # Upload success => update "Uploaded Timestamp" and "ss.ge" columns
                try:
                    wb = openpyxl.load_workbook(excel_path)
                    ws = wb.active

                    # Create col map
                    header = {cell.value: idx for idx, cell in enumerate(ws[1], 1)}

                    ad_id_col = header.get("მესაკუთრის ID")
                    timestamp_col = header.get("Uploaded Timestamp")
                    ss_ge_col = header.get("ss.ge")

                    if not ad_id_col or not timestamp_col or not ss_ge_col:
                        raise ValueError("Missing required columns in Excel (მესაკუთრის ID / Uploaded Timestamp / ss.ge).")

                    current_timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    updated = False

                    for row in ws.iter_rows(min_row=2, values_only=False):
                        cell_ad_id = row[ad_id_col - 1]
                        if cell_ad_id.value == ad_id:
                            timestamp_cell = row[timestamp_col - 1]
                            timestamp_cell.value = current_timestamp

                            ss_ge_cell = row[ss_ge_col - 1]
                            ss_ge_cell.value = final_url

                            updated = True
                            break

                    if not updated:
                        raise ValueError(f"Ad ID {ad_id} not found in Excel for timestamp/URL update.")

                    wb.save(excel_path)
                    logging.info(f"Excel updated with timestamp and final URL for Ad ID: {ad_id}")

                    self.show_info("Scraping completed successfully and data saved to Excel.")
                    self.upload_link.set(f"Upload Successful!\nss.ge: {final_url}")
                    self.copy_button.config(state='normal')

                except Exception as e:
                    logging.error(f"Failed to update Excel with timestamp/URL: {e}")
                    self.show_error(f"Failed to update Excel: {e}")
            else:
                self.show_error("Upload failed. Please check logs for details.")

            # Refresh known IDs
            self.all_ad_ids = self.get_all_ad_ids()

        except Exception as e:
            logging.error(f"An error occurred in run_scrape_upload: {e}")
            self.show_error(f"An error occurred: {e}")
        finally:
            self.progress_scrape_upload.stop()
            self.progress_scrape_upload.pack_forget()
            self.run_button.config(state='normal')
            self.stop_button.pack_forget()
            self.thread_stop_event.clear()

    def start_scrape_only(self):
        if not self.validate_scrape_upload_inputs():
            return
        self.run_button_scrape_only.config(state='disabled')
        self.stop_button_scrape_only.pack(pady=5)
        Thread(target=self.run_scrape_only, daemon=True).start()

    def run_scrape_only(self):
        """
        1) Scrape data
        2) Write minimal fields to Excel, including "SCRAPE ONLY" for timestamp
        3) No upload
        """
        try:
            self.progress_scrape_only.pack(pady=5, fill='x', padx=20)
            self.progress_scrape_only.start()

            headless = self.headless_var_scrape.get()
            if self.thread_stop_event.is_set():
                self.show_info("Process was stopped.")
                return

            data_dir = os.path.join(self.user_data_dir, 'data')
            logging.info(f"Running scraper with data directory: {data_dir}")

            ad_id = run_scraper(
                self.url.get(),
                self.agency_price.get(),
                comment=self.comment.get(),
                headless=headless,
                stop_event=self.thread_stop_event,
                output_dir=data_dir
            )
            logging.info(f"Scraper returned Ad ID: {ad_id}")

            if self.thread_stop_event.is_set():
                self.show_info("Process was stopped.")
                return

            if not ad_id:
                self.show_error("Scraping failed. Check the URL and try again.")
                return

            # Load JSON
            json_file_path = os.path.join(data_dir, ad_id, f"{ad_id}.json")
            if not os.path.exists(json_file_path):
                logging.error("Scraped data file not found.")
                self.show_error("Scraped data file not found.")
                return

            with open(json_file_path, "r", encoding='utf-8') as jf:
                try:
                    scraped_data = json.load(jf)
                    logging.info("Scraped data loaded successfully.")
                except json.JSONDecodeError as e:
                    logging.error(f"JSON decode error: {e}")
                    self.show_error(f"Failed to decode JSON: {e}")
                    return

            flattened_data = flatten_json(scraped_data)

            # We store "SCRAPE ONLY" in Uploaded Timestamp
            excel_data = {
                "Uploaded Timestamp": "SCRAPE ONLY",
                "მესაკუთრის ID": flattened_data.get("ad_id", ""),
                "ტელეფონის ნომერი": flattened_data.get("phone_number", ""),
                "ოთახი": flattened_data.get("property_details_ოთახი", ""),
                "სართული": flattened_data.get("property_details_სართული", ""),
                "მისამართი": f"{flattened_data.get('location', '')} {flattened_data.get('number', '')}".strip(),
                "სა агентოს ფასი": flattened_data.get("agency_price", ""),
                "მესაკუთრის ფასი": flattened_data.get("owner_price", ""),
                "Comment": self.comment.get(),
                "ss.ge": ""
            }

            excel_path = os.path.join(self.user_data_dir, EXCEL_FILE)
            try:
                if not os.path.exists(excel_path):
                    self.ensure_excel_file_exists()

                df = pd.read_excel(excel_path)
                df = pd.concat([df, pd.DataFrame([excel_data])], ignore_index=True)
                df.to_excel(excel_path, index=False)
                logging.info(f"Data appended to Excel for Ad ID: {ad_id}")
            except Exception as e:
                logging.error(f"Failed to write to Excel: {e}")
                self.show_error(f"Failed to write to Excel: {e}")
                return

            self.show_info("Scraping completed successfully and data saved to Excel.")

        except Exception as e:
            logging.error(f"An error occurred in run_scrape_only: {e}")
            self.show_error(f"An error occurred: {e}")
        finally:
            self.progress_scrape_only.stop()
            self.progress_scrape_only.pack_forget()
            self.run_button_scrape_only.config(state='normal')
            self.stop_button_scrape_only.pack_forget()
            self.thread_stop_event.clear()

    def start_upload_existing(self):
        if not self.validate_upload_existing_inputs():
            return
        self.run_button_upload_existing.config(state='disabled')
        self.stop_button_upload_existing.pack(pady=5)
        ad_id = self.existing_ad_id.get()
        Thread(target=self.run_upload_existing, args=(ad_id,), daemon=True).start()

    def run_upload_existing(self, ad_id):
        """
        Re-uploads an existing ad, sets 'ss.ge' if final_url is returned.
        (Optionally you could also set 'Uploaded Timestamp' again if you want.)
        """
        try:
            excel_path = os.path.join(self.user_data_dir, EXCEL_FILE)
            self.progress_upload_existing.pack(pady=5, fill='x', padx=20)
            self.progress_upload_existing.start()
            upload_description = self.upload_description_var_upload.get()

            if self.thread_stop_event.is_set():
                self.show_info("Process was stopped.")
                return

            json_file_path = os.path.join(self.user_data_dir, 'data', ad_id, f"{ad_id}.json")
            logging.info(f"Looking for JSON file at: {json_file_path}")
            if not os.path.exists(json_file_path):
                logging.error("Scraped data file not found.")
                self.show_error("Scraped data file not found.")
                return

            with open(json_file_path, 'r', encoding='utf-8') as jf:
                try:
                    scraped_data = json.load(jf)
                    logging.info("Scraped data loaded successfully.")
                except json.JSONDecodeError as e:
                    logging.error(f"JSON decode error: {e}")
                    self.show_error(f"Failed to decode JSON file: {e}")
                    return

            user_info = self.user_config
            final_url = run_uploader(
                username=user_info['email'],
                password=user_info['password'],
                phone_number=scraped_data.get("phone_number", ""),
                ad_id=ad_id,
                enter_description=upload_description,
                headless=False,
                stop_event=self.thread_stop_event,
                output_dir=os.path.join(self.user_data_dir, 'data')
            )
            logging.info(f"Uploader returned Final URL: {final_url}")

            if self.thread_stop_event.is_set():
                self.show_info("Process was stopped.")
                return

            if final_url:
                try:
                    wb = openpyxl.load_workbook(excel_path)
                    ws = wb.active

                    header = {cell.value: idx for idx, cell in enumerate(ws[1], 1)}
                    ad_id_col = header.get('მესაკუთრის ID')
                    ss_ge_col = header.get('ss.ge')
                    if not ad_id_col:
                        raise ValueError("'მესაკუთრის ID' column not found in Excel.")
                    if not ss_ge_col:
                        ss_ge_col = ws.max_column + 1
                        ws.cell(row=1, column=ss_ge_col, value='ss.ge')

                    updated = False
                    for row in ws.iter_rows(min_row=2, values_only=False):
                        cell_ad_id = row[ad_id_col - 1]
                        if cell_ad_id.value == ad_id:
                            final_url_cell = row[ss_ge_col - 1]
                            final_url_cell.value = final_url
                            updated = True
                            break

                    if not updated:
                        raise ValueError(f"Ad ID {ad_id} not found in Excel for update.")

                    wb.save(excel_path)
                    logging.info(f"ss.ge column updated with final URL for Ad ID: {ad_id}")

                    self.show_info("Scraping completed successfully and data saved to Excel.")
                    self.upload_link.set(f"Upload Successful!\nss.ge: {final_url}")
                    self.copy_button.config(state='normal')

                except Exception as e:
                    logging.error(f"Failed to update ss.ge in Excel: {e}")
                    self.show_error(f"Failed to update ss.ge in Excel: {e}")
            else:
                self.show_error("Upload failed. Check the logs for details.")

        except Exception as e:
            logging.error(f"An error occurred in run_upload_existing: {e}")
            self.show_error(f"An error occurred: {e}")
        finally:
            self.progress_upload_existing.stop()
            self.progress_upload_existing.pack_forget()
            self.run_button_upload_existing.config(state='normal')
            self.stop_button_upload_existing.pack_forget()
            self.thread_stop_event.clear()

    def validate_scrape_upload_inputs(self):
        if not self.url.get() or not self.agency_price.get():
            messagebox.showerror("Input Error", "Please enter both URL and agency price.")
            return False
        return True

    def validate_upload_existing_inputs(self):
        if not self.existing_ad_id.get():
            messagebox.showerror("Input Error", "Please enter an Ad ID.")
            return False
        return True

    def update_ad_id_autocomplete(self, event):
        typed = self.existing_ad_id.get()
        data = []
        if typed == '':
            data = self.all_ad_ids
        else:
            data = [ad_id for ad_id in self.all_ad_ids if typed.lower() in ad_id.lower()]

        if data:
            self.update_ad_id_listbox(data)
            self.ad_id_listbox.pack(pady=5, fill='x', padx=20)
        else:
            self.ad_id_listbox.pack_forget()

    def update_ad_id_listbox(self, data):
        self.ad_id_listbox.delete(0, tk.END)
        for item in data:
            self.ad_id_listbox.insert(tk.END, item)

    def on_ad_id_select(self, event):
        selected_indices = self.ad_id_listbox.curselection()
        if selected_indices:
            index = selected_indices[0]
            selected_ad_id = self.ad_id_listbox.get(index)
            self.existing_ad_id.set(selected_ad_id)
            self.ad_id_listbox.pack_forget()

    def get_all_ad_ids(self):
        data_folder = os.path.join(self.user_data_dir, 'data')
        if not os.path.exists(data_folder):
            return []
        return [name for name in os.listdir(data_folder) if os.path.isdir(os.path.join(data_folder, name))]

    def change_user(self):
        confirm = messagebox.askyesno("Change User", "Are you sure you want to change the user?")
        if confirm:
            config_path = os.path.join(self.user_data_dir, CONFIG_FILE)
            if os.path.exists(config_path):
                try:
                    os.remove(config_path)
                    logging.info("Configuration file deleted.")
                except Exception as e:
                    logging.error(f"Failed to delete config.json: {e}")
                    messagebox.showerror("File Error", f"Failed to delete config.json: {e}")
                    return
            if self.main_frame:
                self.main_frame.pack_forget()
                self.main_frame.destroy()
                self.main_frame = None
            self.show_login_frame()

    def show_error(self, message):
        self.root.after(0, lambda: messagebox.showerror("Error", message))

    def show_info(self, message):
        self.root.after(0, lambda: messagebox.showinfo("Success", message))

    def open_url(self, event):
        url = self.upload_link.get().split("ss.ge: ")[-1]
        if url:
            webbrowser.open(url)

    def copy_url(self):
        url = self.upload_link.get().split("ss.ge: ")[-1]
        if url:
            pyperclip.copy(url)
            messagebox.showinfo("Copied", "URL has been copied to clipboard.")

    def open_excel_file(self):
        excel_path = os.path.join(self.user_data_dir, EXCEL_FILE)
        if not os.path.exists(excel_path):
            self.show_error("Excel file does not exist.")
            return
        try:
            if sys.platform.startswith('darwin'):
                subprocess.call(('open', excel_path))
            elif os.name == 'nt':
                os.startfile(excel_path)
            elif os.name == 'posix':
                subprocess.call(('xdg-open', excel_path))
        except Exception as e:
            logging.error(f"Failed to open Excel file: {e}")
            self.show_error(f"Failed to open Excel file: {e}")

    def stop_running_process(self):
        self.thread_stop_event.set()
        logging.info("Stop event triggered.")

if __name__ == "__main__":
    app = ttk.Window(themename="flatly")
    RealEstateApp(app)
    app.mainloop()
